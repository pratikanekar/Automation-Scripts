import pandas as pd
import subprocess
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

def read_excel_file(file_path):
    try:
        df = pd.read_excel(file_path, header=None, skiprows=1)
        data_list = [value for sublist in df.values.tolist() for value in sublist]
        return data_list
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def check_ping_status(ip):
    try:
        result = subprocess.run(['ping', '-n', '1', ip], capture_output=True, text=True, timeout=5)
        if "Reply from" in result.stdout:
            return "Online"
        else:
            return "Offline"
    except subprocess.TimeoutExpired:
        return "Timeout"

def find_staus(result):
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Serial Number'
        worksheet['B1'] = 'IP Address'
        worksheet['C1'] = 'Ping Status'
        online_count = 0
        offline_count = 0

        for index, ip in enumerate(result, start=2):
            ping_status = check_ping_status(ip)
            worksheet.cell(row=index, column=1, value=index - 1)
            worksheet.cell(row=index, column=2, value=ip)
            worksheet.cell(row=index, column=3, value=ping_status)

            if ping_status == 'Online':
                online_count += 1
            elif ping_status == 'Offline':
                offline_count += 1

            cell = worksheet.cell(row=index, column=3)
            if ping_status == 'Online':
                cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Light green
            elif ping_status == 'Offline':
                cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red

        current_date = datetime.now().strftime("%Y-%m-%d")
        worksheet.cell(row=index + 1, column=1, value='Online Count')
        worksheet.cell(row=index + 1, column=2, value=online_count)
        worksheet.cell(row=index + 2, column=1, value='Offline Count')
        worksheet.cell(row=index + 2, column=2, value=offline_count)
        workbook.save(f"download/router_ping_results_{current_date}.xlsx")
    except Exception as e:
        print(f"{e}")

if __name__ == "__main__":
    file_path = f"upload/router_ip_upload.xlsx"
    print(f"File Uploaded successfully...!")
    result = read_excel_file(file_path)
    print(f"Now analyzing ping status, it will take time. Please wait...")
    find_staus(result)
    print("Successfully downloaded file in the download folder...!")
