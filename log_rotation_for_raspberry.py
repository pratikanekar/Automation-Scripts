from fabric import Connection
from time import sleep
import subprocess
import openpyxl

hostlist = ['10.129.2.39']
username = 'pi'
password = 'iamgw_#@!'
logrotate_conf_content = """# see "man logrotate" for details
# rotate log files weekly
weekly

# keep 4 weeks worth of backlogs
rotate 1

# create new (empty) log files after rotating old ones
create

# use date as a suffix of the rotated file
#dateext

# uncomment this if you want your log files compressed
#compress

# packages drop log rotation information into this directory
include /etc/logrotate.d
"""

rsyslog_conf_content = """/var/log/syslog
{
    rotate 1
    weekly
    missingok
    notifempty
    postrotate
        /usr/lib/rsyslog/rsyslog-rotate
    endscript
}

/var/log/mail.info
/var/log/mail.warn
/var/log/mail.err
/var/log/mail.log
/var/log/daemon.log
/var/log/kern.log
/var/log/auth.log
/var/log/user.log
/var/log/lpr.log
/var/log/cron.log
/var/log/debug
/var/log/messages
{
    rotate 1
    weekly
    missingok
    notifempty
    sharedscripts
    postrotate
        /usr/lib/rsyslog/rsyslog-rotate
    endscript
}
# sharedscripts should be outside the curly braces
sharedscripts
"""

logrotate_command = 'sudo logrotate -f /etc/logrotate.d/rsyslog'
systemctl_content = """[Timer]
OnCalendar=weekly
"""

# Create an Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Ping Status"
sheet.cell(row=1, column=1, value="Hostname")
sheet.cell(row=1, column=2, value="Ping Status")

def is_gateway_pingable(hostname, timeout=2):
    try:
        subprocess.check_output(["ping", "-c", "1", "-W", str(timeout), hostname])
        return True
    except subprocess.CalledProcessError:
        return False

try:
    for index, hostname in enumerate(hostlist, start=2):
        sheet.cell(row=index, column=1, value=hostname)  # Write hostname to Excel
        if is_gateway_pingable(hostname, timeout=2):
            sheet.cell(row=index, column=2, value="True")  # Write True for pingable
            with Connection(
                host=hostname,
                user=username,
                connect_kwargs={"password": password}
            ) as c:
                print(f"Gateway {hostname} is pingable. Configuring logrotate...")

                # Run logrotate configuration commands
                c.sudo(f"echo '{logrotate_conf_content}' | sudo tee /etc/logrotate.conf > /dev/null")
                c.sudo(f"echo '{rsyslog_conf_content}' | sudo tee /etc/logrotate.d/rsyslog > /dev/null")
                c.sudo(f"{logrotate_command} > /dev/null")

                print(f"Configuring logrotate.timer on {hostname}")
                # Creating the directory if it doesn't exist
                c.sudo("mkdir -p /etc/systemd/system/logrotate.timer.d")
                c.sudo(f"echo '{systemctl_content}' | sudo tee /etc/systemd/system/logrotate.timer.d/override.conf > /dev/null")
                c.sudo("sudo systemctl daemon-reload > /dev/null")
                c.sudo("sudo systemctl restart logrotate.timer > /dev/null")

                print(f"Configuration completed successfully on {hostname}")
        else:
            sheet.cell(row=index, column=2, value="False")  # Write False for not pingable
            print(f"Gateway {hostname} is not pingable. Skipping configuration.")
        sleep(0.5)

except Exception as e:
    print(f"Configuration failed on {hostname}. Error: {e}")

# Save the Excel file
workbook.save(f"download/log_rotate_ping_status.xlsx")
print("file was created successfully...")
